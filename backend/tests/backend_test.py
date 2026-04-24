"""Backend API tests for Insurance Benefits Summarizer."""
import io
import time
import requests
import pytest
from docx import Document as DocxDocument
from openpyxl import Workbook, load_workbook


# ---------- Root / Health ----------
class TestRoot:
    def test_root(self, base_url):
        r = requests.get(f"{base_url}/api/", timeout=20)
        assert r.status_code == 200
        data = r.json()
        assert data.get("status") == "ok"


# ---------- Auth ----------
class TestAuth:
    def test_wrong_pin_returns_401(self, base_url):
        r = requests.post(f"{base_url}/api/auth/pin", json={"pin": "0000"}, timeout=20)
        assert r.status_code == 401

    def test_correct_pin_returns_jwt(self, base_url, app_pin):
        r = requests.post(f"{base_url}/api/auth/pin", json={"pin": app_pin}, timeout=20)
        assert r.status_code == 200
        data = r.json()
        assert "token" in data and isinstance(data["token"], str) and len(data["token"]) > 20
        assert "expires_at" in data

    def test_me_rejects_no_token(self, base_url):
        r = requests.get(f"{base_url}/api/auth/me", timeout=20)
        assert r.status_code in (401, 403)

    def test_me_rejects_invalid_token(self, base_url):
        r = requests.get(
            f"{base_url}/api/auth/me",
            headers={"Authorization": "Bearer not-a-real-token"},
            timeout=20,
        )
        assert r.status_code == 401

    def test_me_accepts_valid_token(self, base_url, auth_headers):
        r = requests.get(f"{base_url}/api/auth/me", headers=auth_headers, timeout=20)
        assert r.status_code == 200
        data = r.json()
        assert data.get("authenticated") is True


# ---------- Analyses list auth ----------
class TestAnalysesAuth:
    def test_list_requires_auth(self, base_url):
        r = requests.get(f"{base_url}/api/analyses", timeout=20)
        assert r.status_code in (401, 403)

    def test_list_with_auth(self, base_url, auth_headers):
        r = requests.get(f"{base_url}/api/analyses", headers=auth_headers, timeout=20)
        assert r.status_code == 200
        assert isinstance(r.json(), list)


# ---------- Analyze (uses real LLM) ----------
SAMPLE_TEXT = (
    "ACME Corp 2025 Benefits Summary\n"
    "Carrier: Guardian Life. Effective Date: 01/01/2025.\n"
    "Dental PPO: In-network deductible $50 individual / $150 family; coinsurance Preventive 100%, "
    "Basic 80%, Major 50%; annual max $1500.\n"
    "Dental rates: EE $28.50/month, EE+Spouse $58.00/month, EE+Family $92.75/month.\n"
    "Vision: Exam $10 copay every 12 months; Frames allowance $150; Contacts allowance $150.\n"
    "Vision rates: EE $6.25/month, Family $18.40/month.\n"
    "Basic Life: 1x annual salary up to $50,000 guaranteed issue. AD&D matching.\n"
    "Voluntary Life: Employee up to $500,000, Spouse up to $250,000, Child $10,000.\n"
    "STD: 60% of weekly earnings, 7 day elimination, 26 weeks benefit duration.\n"
    "LTD: 60% of monthly earnings to $10,000 max, 90 day elimination, to SSNRA.\n"
)


def _make_docx_bytes(text: str) -> bytes:
    doc = DocxDocument()
    for line in text.split("\n"):
        doc.add_paragraph(line)
    buf = io.BytesIO()
    doc.save(buf)
    buf.seek(0)
    return buf.read()


def _make_xlsx_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Rates"
    ws.append(["Tier", "Rate"])
    ws.append(["EE", "$28.50"])
    ws.append(["Family", "$92.75"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


class TestAnalyze:
    def test_analyze_requires_auth(self, base_url):
        files = {"files": ("sample.txt", b"hello", "text/plain")}
        r = requests.post(f"{base_url}/api/analyze", files=files, timeout=30)
        assert r.status_code in (401, 403)

    @pytest.mark.timeout(180)
    def test_analyze_multi_format_success(self, base_url, auth_headers, request):
        docx_bytes = _make_docx_bytes(SAMPLE_TEXT)
        xlsx_bytes = _make_xlsx_bytes()
        files = [
            ("files", ("benefits.docx", docx_bytes,
                       "application/vnd.openxmlformats-officedocument.wordprocessingml.document")),
            ("files", ("rates.xlsx", xlsx_bytes,
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ]
        # multipart -> don't force JSON header
        r = requests.post(
            f"{base_url}/api/analyze",
            headers=auth_headers,
            files=files,
            timeout=180,
        )
        assert r.status_code == 200, f"Analyze failed: {r.status_code} {r.text[:500]}"
        data = r.json()
        # Required keys
        for k in ["id", "title", "created_at", "source_files", "executive_summary",
                  "carriers", "effective_date", "benefits"]:
            assert k in data, f"missing key {k}"
        assert "_id" not in data
        # Benefit keys
        for k in ["dental", "vision", "basic_life", "voluntary_life", "std", "ltd"]:
            assert k in data["benefits"], f"missing benefit line {k}"
            b = data["benefits"][k]
            assert "plan_design" in b and "rates" in b and "coverages" in b and "notes" in b
        assert len(data["source_files"]) == 2
        # Stash id for later tests
        request.config.cache.set("analysis_id", data["id"])


# ---------- CRUD ----------
class TestAnalysisCRUD:
    def test_get_by_id(self, base_url, auth_headers, request):
        analysis_id = request.config.cache.get("analysis_id", None)
        if not analysis_id:
            pytest.skip("No analysis_id from previous test")
        r = requests.get(f"{base_url}/api/analyses/{analysis_id}",
                         headers=auth_headers, timeout=20)
        assert r.status_code == 200
        data = r.json()
        assert data["id"] == analysis_id
        assert "_id" not in data

    def test_list_contains_id(self, base_url, auth_headers, request):
        analysis_id = request.config.cache.get("analysis_id", None)
        if not analysis_id:
            pytest.skip("No analysis_id from previous test")
        r = requests.get(f"{base_url}/api/analyses", headers=auth_headers, timeout=20)
        assert r.status_code == 200
        rows = r.json()
        ids = [row["id"] for row in rows]
        assert analysis_id in ids
        for row in rows:
            assert "_id" not in row

    def test_export_requires_token(self, base_url, request):
        analysis_id = request.config.cache.get("analysis_id", None)
        if not analysis_id:
            pytest.skip("No analysis_id from previous test")
        r = requests.get(f"{base_url}/api/analyses/{analysis_id}/export", timeout=30)
        assert r.status_code == 401

    def test_export_with_token(self, base_url, auth_token, request):
        analysis_id = request.config.cache.get("analysis_id", None)
        if not analysis_id:
            pytest.skip("No analysis_id from previous test")
        r = requests.get(
            f"{base_url}/api/analyses/{analysis_id}/export",
            params={"token": auth_token},
            timeout=30,
        )
        assert r.status_code == 200
        assert "spreadsheet" in r.headers.get("content-type", "")
        # Validate bytes are a valid xlsx
        wb = load_workbook(io.BytesIO(r.content))
        assert "Overview" in wb.sheetnames

    def test_delete_and_verify(self, base_url, auth_headers, request):
        analysis_id = request.config.cache.get("analysis_id", None)
        if not analysis_id:
            pytest.skip("No analysis_id from previous test")
        r = requests.delete(f"{base_url}/api/analyses/{analysis_id}",
                            headers=auth_headers, timeout=20)
        assert r.status_code == 200
        assert r.json().get("deleted") is True

        # Verify 404 on subsequent GET
        r2 = requests.get(f"{base_url}/api/analyses/{analysis_id}",
                          headers=auth_headers, timeout=20)
        assert r2.status_code == 404

    def test_delete_nonexistent_404(self, base_url, auth_headers):
        r = requests.delete(f"{base_url}/api/analyses/does-not-exist",
                            headers=auth_headers, timeout=20)
        assert r.status_code == 404
